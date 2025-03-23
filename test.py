"""Создание файлов и папок"""


import os
import openpyxl
import shutil
import locale

from pathlib import Path

# Устанавливаем локаль
locale.setlocale(locale.LC_TIME, "ru_RU.UTF-8")

# Определяем пути
parent_dir = Path("w")
workbook = openpyxl.load_workbook(parent_dir / "Заявки ВОЛС 2022 - 2025.xlsx")
sheet = workbook["2025"]


def create_folders(operator, date_folder, folder_name):
    """Создает необходимые папки"""
    base_path = parent_dir / operator / "АВР_ДР" / \
        "на оформлении" / date_folder / folder_name
    (base_path / "Фото").mkdir(parents=True, exist_ok=True)
    return base_path


def copy_and_edit_file(template_path, destination_folder, folder_name, replacements):
    """Копирует и редактирует файл"""
    destination_file = destination_folder / f"{folder_name}.xlsx"
    shutil.copy2(template_path, destination_file)

    wb = openpyxl.load_workbook(destination_file)
    for sheet_name, updates in replacements.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for cell, value in updates.items():
                ws[cell] = value
        else:
            print(
                f"Ошибка: Лист '{sheet_name}' не найден. Доступные листы: {wb.sheetnames}")
            wb.close()
            return
    wb.save(destination_file)
    wb.close()


# Обработка строк из Excel
for row in sheet.iter_rows(min_row=45, max_row=56, min_col=1, max_col=sheet.max_column):
    number, operator = row[0].value, row[1].value
    date_obj = row[3].value
    date, time = date_obj.strftime("%d.%m.%Y"), date_obj.strftime("%H:%M")
    date_folder, date_full = date_obj.strftime(
        "%m_%Y"), date_obj.strftime("%d %B %Y")
    folder_name = f"{row[2].value} {date} {row[4].value} {row[6].value} {row[5].value}"

    # Создаем папки
    destination_folder = create_folders(operator, date_folder, folder_name)

    # Определяем шаблон и редактируемый контент
    template_files = {
        "Мегафон": "w/оператор_бланк/megafon.xlsx",
        "ВПК гор": "w/оператор_бланк/beeline.xlsx"
    }
    if operator in template_files:
        template_path = Path(template_files[operator])
        replacements = {
            "ВедомостьВР": {"C10": f"{row[2].value} от {date}"},
            "11Акт": {"B2": f" от {date_full}", "B7": f"1. {date} г. Исполнитель \
                      оказал, а Заказчик принял работы по договору \
                      № РТК-ВОЛС-24-26 от 15.03.2024 г., а именно:"},
            "Акт ДР": {"E4": date_full, "D15": date_full, "D17": date_full, "E15": time}
        }
        copy_and_edit_file(template_path, destination_folder,
                           folder_name, replacements)

workbook.close()
