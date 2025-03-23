"""Создание файлов и папок"""


import os

import openpyxl
import shutil
import locale

from pathlib import Path


locale.setlocale(locale.LC_TIME, "ru_RU.UTF-8")
parent_dir = Path("w")
workbook = openpyxl.load_workbook("w/Заявки ВОЛС 2022 - 2025.xlsx")
sheet = workbook["2025"]


# Проходим по всем строкам и столбцам
for row in sheet.iter_rows(min_row=45, max_row=56, min_col=1, max_col=sheet.max_column):
    number = row[0].value
    date = row[3].value.strftime("%d.%m.%Y")
    time = row[3].value.strftime("%H:%M")
    date_folder = row[3].value.strftime("%m_%Y")
    date_full = row[3].value.strftime("%d %B %Y")
    folder_name = (str(row[2].value)+' '+str(date)+' ' +
                   str(row[4].value)+' '+str(row[6].value)+' '+str(row[5].value))
    operator = row[1].value
    child_dir = parent_dir / operator / "АВР_ДР" / \
        "на оформлении" / date_folder / folder_name
    child_dir.mkdir(parents=True, exist_ok=True)
    child_photo = parent_dir / operator / "АВР_ДР" / \
        "на оформлении" / date_folder / folder_name / "Фото"
    child_photo.mkdir(parents=True, exist_ok=True)

    """Megafon"""
    if (operator == 'Мегафон'):
        # Исходный файл и путь назначения
        source_file = "w/оператор_бланк/megafon.xlsx"  # Укажите путь к исходному файлу
        destination_folder = child_dir  # Укажите путь к папке назначения
        new_filename = f"{folder_name}.xlsx"  # Новое имя файла
        # Проверяем, существует ли папка назначения
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)
        # Полный путь к новому файлу
        destination_file = os.path.join(destination_folder, new_filename)
        # Копируем и переименовываем файл
        shutil.copy2(source_file, destination_file)
        # Открываем файл для редактирования
        wb = openpyxl.load_workbook(destination_file)
        # Выбираем нужный лист (например, "ВедомостьВР")
        sheet_name = "ВедомостьВР"  # Название листа
        sheet_name_akt = "11Акт"  # Название листа
        sheet_name_dr = "Акт ДР"  # Название листа
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_akt = wb[sheet_name_akt]
            ws_dr = wb[sheet_name_dr]
        else:
            print(
                f"Ошибка: Лист '{sheet_name}' не найден. Доступные листы: {wb.sheetnames}")
            print(
                f"Ошибка: Лист '{sheet_name_akt}' не найден. Доступные листы: {wb.sheetnames}")
            print(
                f"Ошибка: Лист '{sheet_name_dr}' не найден. Доступные листы: {wb.sheetnames}")
            wb.close()
            exit()
        # Вносим изменения
        ws["C10"] = str(row[2].value) + " от " + date
        workbook_tg = openpyxl.load_workbook("w/list.xlsx")
        tg = workbook_tg["Лист1"]
        for tg in tg.iter_rows(min_row=1, max_row=tg.max_row, min_col=1, max_col=tg.max_column):
            if (str(tg[0].value) == str(number)):
                ws["D45"] = tg[11].value * 2
            workbook_tg.close()
        ws_akt["B2"] = " от " + date_full
        ws_akt["B7"] = f"1. {date} г. Исполнитель оказал, а Заказчик принял работы по договору № РТК-ВОЛС-24-26 от 15.03.2024 г., а именно:"
        ws_dr["E4"] = ws_dr["D15"] = ws_dr["D17"] = date_full
        ws_dr["E15"] = time
        # Сохраняем файл
        wb.save(destination_file)
        wb.close()

    """ВПК гор"""
    if (operator == 'ВПК гор'):
        # Исходный файл и путь назначения
        source_file = "w/оператор_бланк/beeline.xlsx"  # Укажите путь к исходному файлу
        destination_folder = child_dir  # Укажите путь к папке назначения
        new_filename = f"{folder_name}.xlsx"  # Новое имя файла
        # Проверяем, существует ли папка назначения
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)
        # Полный путь к новому файлу
        destination_file = os.path.join(destination_folder, new_filename)
        # Копируем и переименовываем файл
        shutil.copy2(source_file, destination_file)
        # Открываем файл для редактирования
        wb = openpyxl.load_workbook(destination_file)
        # Выбираем нужный лист (например, "ВедомостьВР")

        sheet_name = "ВедомостьВР"  # Название листа
        sheet_name_akt = "1 Акт"  # Название листа
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            print(
                f"Ошибка: Лист '{sheet_name}' не найден. Доступные листы: {wb.sheetnames}")
            wb.close()
            exit()
        # Вносим изменения
        ws["C5"] = str(row[2].value) + " от " + date
        workbook_tg = openpyxl.load_workbook("w/list.xlsx")
        tg = workbook_tg["Лист1"]
        for tg in tg.iter_rows(min_row=1, max_row=tg.max_row, min_col=1, max_col=tg.max_column):
            if (str(tg[0].value) == str(number)):
                ws["D31"] = tg[11].value * 2
        workbook_tg.close()
        # Сохраняем файл
        wb.save(destination_file)
        wb.close()

workbook.close()
